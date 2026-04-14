"""
This module provides functions and utilities for exporting tables from a DexpiModel
to various formats (Excel, CSV, JSON). It includes logic for formatting, column ordering,
and attribute extraction for equipment, lines, instruments, and piping components.
"""

import json
import re
from collections.abc import Iterable
from enum import Enum, StrEnum
from pathlib import Path
from typing import Literal

import pandas as pd
from openpyxl import load_workbook

from pydexpi.dexpi_classes import equipment, instrumentation, physicalQuantities, piping
from pydexpi.dexpi_classes.dexpiBaseModels import DexpiBaseModel
from pydexpi.dexpi_classes.dexpiModel import DexpiModel
from pydexpi.toolkits.base_model_utils import get_composition_attributes, get_data_attributes
from pydexpi.toolkits.model_toolkit import get_all_instances_in_model
from pydexpi.toolkits.piping_toolkit import add_system_attributes_to_segment

COLUMN_PRIORITIES = {
    "equipment": ["Tag", "Equipment type"],
    "line": ["Segment number"],
    "instrument": ["Tag", "Operated valve", "Sensing location"],
    "piping component": ["Piping component number", "Piping component"],
}


class ListType(StrEnum):
    """Enum for the different list types that can be exported."""

    EQUIPMENT = "equipment"
    LINE = "line"
    INSTRUMENT = "instrument"
    PIPING_COMPONENT = "piping component"
    ALL = "all"


class ExportFormat(StrEnum):
    """Enum for the different export formats."""

    XLSX = ".xlsx"
    CSV = ".csv"
    JSON = ".json"


def set_excel_column_widths(excel_path: str, tables: dict) -> None:
    """
    Set Excel column widths for each sheet based on header length. For the first column, make sure
    that all content is visible.

    Parameters
    ----------
    excel_path: str
        The path to the Excel file to modify.
    tables: dict
        The mapping of {sheet_name: DataFrame}

    Returns
    -------
    None
    """
    wb = load_workbook(excel_path)
    for sheet_name, df in tables.items():
        ws = wb[sheet_name]
        for idx, col in enumerate(df.columns, 1):
            if idx == 1:
                # First column: set width based on max content length or header length
                max_length = max(len(str(cell)) for cell in df[col].astype(str).tolist() + [col])
                ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = (
                    max_length + 2
                )
            else:
                # Other columns: set width based on header length
                header_length = len(str(col))
                ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = (
                    header_length + 2
                )
    wb.save(excel_path)


def export_tables(
    model: DexpiModel,
    path_stem: str,
    lists: Iterable[Literal["all", "equipment", "line", "instrument", "piping component"]]
    | None = None,
    formats: Iterable[Literal[".xlsx", ".csv", ".json"]] | None = None,
) -> None:
    """
    Main function to build and export selected lists from a DexpiModel.

    Parameters
    ----------
    model : DexpiModel
        The DEXPI model to export tables from.
    path_stem : str
        The path and filename stem for the output files (without extension).
    lists : Iterable of Literal {"all", "equipment", "line", "instrument", "piping component"}, optional
        The types of lists to export. If ``None``, all lists are exported.
    formats : Iterable of Literal {".xlsx", ".csv", ".json"}, optional
        The formats to export the lists in. If ``None``, defaults to ``.xlsx``.

    Returns
    -------
    None
        Files are written to disk at paths derived from ``path_stem``.
    """
    path_stem = Path(path_stem)
    if lists is None:
        lists = (ListType.ALL,)
    if formats is None:
        formats = (ExportFormat.XLSX,)

    # Prepare all tables
    tables = {
        ListType.EQUIPMENT: make_equipment_list(model),
        ListType.LINE: make_line_list(model),
        ListType.INSTRUMENT: make_instrument_list(model),
        ListType.PIPING_COMPONENT: make_piping_component_list(model),
    }

    # Clean up tables (drop None columns, reorder)
    for key, df in tables.items():
        df = df.dropna(axis=1, how="all")
        df = reorder_columns(df, COLUMN_PRIORITIES.get(key.value, []))
        tables[key] = df

    for format in formats:
        format = ExportFormat(format)
        if ListType.ALL in lists:
            # Export all tables in one file
            if format == ExportFormat.XLSX:
                out_path = path_stem.with_suffix(format)
                with pd.ExcelWriter(out_path) as writer:
                    for key, df in tables.items():
                        df.to_excel(writer, sheet_name=key.value.capitalize(), index=False)
                # Set column widths for each sheet
                set_excel_column_widths(
                    out_path, {key.value.capitalize(): df for key, df in tables.items()}
                )
            elif format == ExportFormat.JSON:
                out_path = path_stem.with_suffix(format)
                tables_dict = {
                    key.value: df.where(pd.notnull(df), None).to_dict(orient="records")
                    for key, df in tables.items()
                }
                with open(out_path, "w", encoding="utf-8") as f:
                    json.dump(tables_dict, f, indent=4)
            elif format == ExportFormat.CSV:
                # Export each table as a separate CSV file
                for key, df in tables.items():
                    out_path = path_stem.with_name(f"{path_stem.name}_{key.value}").with_suffix(
                        format
                    )
                    df.to_csv(out_path, index=False)
        else:
            # Export selected tables individually
            for list_type in lists:
                key = ListType(list_type)
                df = tables[key]
                out_path = path_stem.with_name(f"{path_stem.name}_{key.value}").with_suffix(format)
                if format == ExportFormat.XLSX:
                    df.to_excel(out_path, index=False)
                    set_excel_column_widths(out_path, {key.value.capitalize(): df})
                elif format == ExportFormat.JSON:
                    df = df.where(pd.notnull(df), None)
                    df.to_json(out_path, orient="records", indent=4)
                elif format == ExportFormat.CSV:
                    df.to_csv(out_path, index=False)


def reorder_columns(df: pd.DataFrame, priority: list[str]) -> pd.DataFrame:
    """Reorder the columns of a DataFrame based on a given priority list. Sort rest alphabetically.

    Parameters
    ----------
    df : pd.DataFrame
        The DataFrame to reorder.
    priority : list[str]
        The list of column names to prioritize in the order they should appear.

    Returns
    -------
    pd.DataFrame
        The DataFrame with reordered columns.
    """
    cols = df.columns.tolist()
    ordered = [c for c in priority if c in cols]
    remaining = sorted([c for c in cols if c not in ordered])
    return df[ordered + remaining]


def format_string(s: str) -> str:
    """Format pyDEXPI attribute names for display.

    Parameters
    ----------
    s : str
        The original attribute name in camelCase or PascalCase.

    Returns
    -------
    str
        A human-readable label with the first character capitalized and spaces
        inserted before internal capital letters (which are lowercased).
    """
    # Capitalize the first character
    first = s[0].upper()
    # Add space before every other capital letter and make it lowercase
    rest = re.sub(r"([A-Z])", lambda m: " " + m.group(1).lower(), s[1:])
    # Combine first character with the rest
    return first + rest


def float_to_str(value: float) -> str:
    """Convert a float to a compact string.

    Parameters
    ----------
    value : float
        The numeric value to convert.

    Returns
    -------
    str
        ``"<int>"`` if ``value`` is a whole number (e.g., ``3.0 -> "3"``), otherwise
        the standard string representation of the float.
    """
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def get_attribute_value(value_object: DexpiBaseModel) -> str | None:
    """Resolve a DEXPI data attribute object to a printable string.

    Parameters
    ----------
    value_object : DexpiBaseModel
        The attribute value, which may be a physical quantity, Enum, or str.

    Returns
    -------
    str | None
        A string formatted as "<value> <unit>" for physical quantities, the
        underlying ``.value`` for Enums, the original string for ``str`` inputs,
        or ``None`` if the type is not handled.
    """
    if hasattr(physicalQuantities, value_object.__class__.__name__):
        return f"{float_to_str(value_object.value)} {value_object.unit.value}"
    elif isinstance(value_object, Enum):
        return value_object.value
    elif isinstance(value_object, str):
        return value_object
    return None


def extend_row_with_subequipment(row: dict, comp_object: DexpiBaseModel, i: int | None) -> dict:
    """Extend a parent equipment row with sub-equipment data attributes.

    Parameters
    ----------
    row : dict
        The row under construction for the parent equipment.
    comp_object : DexpiBaseModel
        The sub-equipment object whose data attributes will be appended.
    i : int | None
        Index of the sub-equipment within a list (used to disambiguate names),
        or ``None`` if the sub-equipment is a single object.

    Returns
    -------
    dict
        The updated row containing additional columns for the sub-equipment's
        formatted data attributes.
    """
    # Get the sub equipment tag
    if hasattr(comp_object, "subTagName"):
        sub_tag = comp_object.subTagName
    else:
        sub_tag = None
    if sub_tag is None:
        sub_tag = comp_object.__class__.__name__
        if i is not None:
            sub_tag += " " + str(i + 1)

    # Get sub equipment data attributes
    data_comp_attributes = get_data_attributes(comp_object)
    for (
        data_comp_name,
        data_comp_value_object,
    ) in data_comp_attributes.items():
        if data_comp_name == "subTagName":
            continue
        formated_data_comp_name = sub_tag + ", " + format_string(data_comp_name)
        sub_value = get_attribute_value(data_comp_value_object)
        row[formated_data_comp_name] = sub_value

    return row


def make_equipment_list(model: DexpiModel) -> pd.DataFrame:
    """Collect all equipment instances and their attributes as a table.

    Parameters
    ----------
    model : DexpiModel
        The DEXPI model containing equipment instances.

    Returns
    -------
    pd.DataFrame
        A table where each row represents a tagged plant item and columns include
        identifiers, data attributes, and flattened sub-equipment attributes.
    """

    table = []
    tagged_plant_items = get_all_instances_in_model(model, equipment.TaggedPlantItem)
    for item in tagged_plant_items:
        row = {}
        # Identifier
        row["Tag"] = item.tagName
        row["Equipment type"] = item.__class__.__name__
        # data attributes
        data_attributes = get_data_attributes(item)
        for name, value_object in data_attributes.items():
            if "tag" not in name.lower():  # skip tagName again
                formated_name = format_string(name)
                value = get_attribute_value(value_object)
                row[formated_name] = value
        # compositional attributes -> sub equipment
        comp_attributes = get_composition_attributes(item)
        for comp_objects in comp_attributes.values():
            # data attributes of sub equipment
            if isinstance(comp_objects, list):
                for i, comp_object in enumerate(comp_objects):
                    row = extend_row_with_subequipment(row, comp_object, i)
            else:
                row = extend_row_with_subequipment(row, comp_objects, None)

        table.append(row)

    equipment_list = pd.DataFrame(table)
    return equipment_list


def make_line_list(model: DexpiModel) -> pd.DataFrame:
    """Collect piping network segments and their attributes as a table.

    Parameters
    ----------
    model : DexpiModel
        The DEXPI model containing piping network systems and segments.

    Returns
    -------
    pd.DataFrame
        A table with one row per piping network segment including its data attributes
        (with system-level attributes propagated to segments where applicable).
    """

    # TODO Line identifier currently limited, line tags not parsed from Proteus yet. Use segment number for now.
    table = []
    systems = get_all_instances_in_model(model, piping.PipingNetworkSystem)
    for system in systems:
        system = add_system_attributes_to_segment(system)
        for segment in system.segments:
            row = {}
            # Data attributes
            data_attributes = get_data_attributes(segment)
            for name, value_object in data_attributes.items():
                formated_name = format_string(name)
                value = get_attribute_value(value_object)
                row[formated_name] = value
            table.append(row)

    line_list = pd.DataFrame(table)
    return line_list


def make_instrument_list(model: DexpiModel) -> pd.DataFrame:
    """Collect process instrumentation functions and their attributes as a table.

    Parameters
    ----------
    model : DexpiModel
        The DEXPI model containing instrumentation functions and related objects.

    Returns
    -------
    pd.DataFrame
        A table with one row per instrumentation function including Tag, operated
        valve (if any), sensing location, and other data attributes.
    """
    table = []
    instruments = get_all_instances_in_model(model, instrumentation.ProcessInstrumentationFunction)
    for instrument_object in instruments:
        row = {}
        # Identifier
        row["Tag"] = (
            instrument_object.processInstrumentationFunctionCategory
            + instrument_object.processInstrumentationFunctions
            + "-"
            + instrument_object.processInstrumentationFunctionNumber
        )
        # Operated valve reference
        actuating_functions = instrument_object.actuatingFunctions
        for actuating_function in actuating_functions:
            actuating_system = actuating_function.systems
            if actuating_system:
                operatedValveReference = actuating_system.operatedValveReference
                if operatedValveReference:
                    row["Operated valve"] = operatedValveReference.valve.pipingComponentNumber
        # Sensing location
        generating_functions = instrument_object.processSignalGeneratingFunctions
        for generating_function in generating_functions:
            sensing_location = generating_function.sensingLocation
            if sensing_location:
                if issubclass(sensing_location.__class__, piping.PipingComponent):
                    row["Sensing location"] = (
                        sensing_location.pipingComponentNumber
                    )  # TODO replace with tag
                elif issubclass(sensing_location.__class__, piping.PipingNetworkSegment):
                    row["Sensing location"] = (
                        sensing_location.segmentNumber
                    )  # TODO replace with tag
                elif issubclass(sensing_location.__class__, equipment.Nozzle):
                    row["Sensing location"] = (
                        sensing_location.subTagName
                    )  # TODO check if this makes sense
        # Data attributes
        data_attributes = get_data_attributes(instrument_object)
        for name, value_object in data_attributes.items():
            formated_name = format_string(name)
            value = get_attribute_value(value_object)
            row[formated_name] = value
        table.append(row)

    instrument_list = pd.DataFrame(table)
    return instrument_list


def make_piping_component_list(model: DexpiModel) -> pd.DataFrame:
    """Collect piping components and their attributes as a table.

    Parameters
    ----------
    model : DexpiModel
        The DEXPI model containing piping components.

    Returns
    -------
    pd.DataFrame
        A table with one row per piping component including its class name and data
        attributes.
    """

    # TODO Valve tag is currently not parsed from Proteus. Use PipingComponentNumber for now.
    table = []
    piping_components = get_all_instances_in_model(model, piping.PipingComponent)
    for piping_component_object in piping_components:
        row = {}
        row["Piping component"] = piping_component_object.__class__.__name__
        # Data attributes
        data_attributes = get_data_attributes(piping_component_object)
        for name, value_object in data_attributes.items():
            formated_name = format_string(name)
            value = get_attribute_value(value_object)
            row[formated_name] = value
        table.append(row)

    piping_component_list = pd.DataFrame(table)
    return piping_component_list
